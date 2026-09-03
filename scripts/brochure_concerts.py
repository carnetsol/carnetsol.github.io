#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
brochure_concerts.py — Génère une brochure PDF à partir d'un relevé de concerts .xlsx
(Agenda de Carnets sur sol)

USAGE
-----
    python3 brochure_concerts.py tous_concerts.xlsx
    python3 brochure_concerts.py tous_concerts.xlsx -o tous_concerts.pdf --logo cello.jpg
    python3 brochure_concerts.py tous_concerts.xlsx --feuille "À venir" --du 2026-09-01 --au 2027-08-31

STRUCTURE DU FICHIER XLSX ATTENDUE
----------------------------------
Ligne 1 = en-têtes. Colonnes, dans cet ordre :
    A Jour  (format « 2026-09-01 (Mardi) »)   B Heure    C Salle
    D Compositeur(s) & Œuvre(s)                E Interprètes
    F Commentaires                             G Type      H Tarifs

CODES COULEUR (remplissage de cellule)
--------------------------------------
    rose  #E0C2CD  sur la colonne D (Œuvres)  → « hautement signalé » (bordeaux)
    rose  ailleurs sur la ligne               → « signalé » (vert)
    vert  #F6F9D4  n'importe où sur la ligne  → « signalé » (vert)
Les teintes se règlent via --rose et --vert si votre palette diffère.

DÉPENDANCES
-----------
    pip install openpyxl reportlab
Polices : DejaVu Serif / Sans (paquet fonts-dejavu sous Debian/Ubuntu).
Sur macOS/Windows, passez --polices vers un dossier contenant les .ttf DejaVu,
sinon le script bascule sur Helvetica/Times (les diacritiques tchèques,
polonais, hongrois risquent alors de mal s'afficher).
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta

# ───────────────────────────────────────────────────────────────────────
#  Dépendances externes
# ───────────────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("Module manquant : openpyxl.  →  pip install openpyxl")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, HRFlowable,
        PageBreak, Table, TableStyle, Image,
    )
except ImportError:
    sys.exit("Module manquant : reportlab.  →  pip install reportlab")


# ───────────────────────────────────────────────────────────────────────
#  Polices
# ───────────────────────────────────────────────────────────────────────
DEJAVU_DIRS = [
    '/usr/share/fonts/truetype/dejavu/',
    '/usr/share/fonts/dejavu/',
    '/usr/local/share/fonts/dejavu/',
    '/Library/Fonts/',
    os.path.expanduser('~/Library/Fonts/'),
    'C:/Windows/Fonts/',
]

# nom logique → (fichier DejaVu, repli intégré à reportlab)
FONT_FILES = {
    'SERIF':      ('DejaVuSerif.ttf',           'Times-Roman'),
    'SERIF-B':    ('DejaVuSerif-Bold.ttf',      'Times-Bold'),
    'SERIF-I':    ('DejaVuSerif-Italic.ttf',    'Times-Italic'),
    'SANS':       ('DejaVuSans.ttf',            'Helvetica'),
    'SANS-B':     ('DejaVuSans-Bold.ttf',       'Helvetica-Bold'),
}

FONT = {}   # nom logique → nom de police réellement enregistré


def setup_fonts(extra_dir=None):
    """Enregistre DejaVu si disponible, sinon bascule sur les polices de base."""
    search = ([extra_dir] if extra_dir else []) + DEJAVU_DIRS
    found_dir = None
    for d in search:
        if d and os.path.isfile(os.path.join(d, 'DejaVuSerif.ttf')):
            found_dir = d
            break

    if found_dir:
        for logical, (fname, _fallback) in FONT_FILES.items():
            path = os.path.join(found_dir, fname)
            reg_name = 'DJ' + logical
            try:
                pdfmetrics.registerFont(TTFont(reg_name, path))
                FONT[logical] = reg_name
            except Exception:
                FONT[logical] = FONT_FILES[logical][1]
    else:
        print("! Polices DejaVu introuvables — repli sur Helvetica/Times.\n"
              "  Certains caractères accentués rares peuvent mal s'afficher.\n"
              "  (option --polices pour indiquer un dossier de .ttf DejaVu)",
              file=sys.stderr)
        for logical, (_f, fallback) in FONT_FILES.items():
            FONT[logical] = fallback


# ───────────────────────────────────────────────────────────────────────
#  Palette
# ───────────────────────────────────────────────────────────────────────
CREME    = colors.HexColor('#F7F3EB')   # fond courant
CREME_D  = colors.HexColor('#EDE8DC')   # fond des pages de mois
GRIS_F   = colors.HexColor('#232320')
GRIS_M   = colors.HexColor('#6B6660')
GRIS_L   = colors.HexColor('#C8C0B0')
OCRE     = colors.HexColor('#B07A20')
OCRE_L   = colors.HexColor('#F0E4C8')
BORD     = colors.HexColor('#7A1F3A')   # hautement signalé
BORD_L   = colors.HexColor('#F5EAEF')
VERT     = colors.HexColor('#2D5E40')   # signalé
VERT_L   = colors.HexColor('#E8F2EC')

W, H = A4
ML = MR = 23 * mm
MT = MB = 18 * mm

MOIS_FR = {1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
           5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
           9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'}
JOURS_FR = {0: 'Lundi', 1: 'Mardi', 2: 'Mercredi', 3: 'Jeudi',
            4: 'Vendredi', 5: 'Samedi', 6: 'Dimanche'}


# ───────────────────────────────────────────────────────────────────────
#  Dates
# ───────────────────────────────────────────────────────────────────────
def parse_date(s):
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', str(s))
    if not m:
        return None
    try:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
    except ValueError:
        return None


def week_start(d):
    """Lundi de la semaine contenant d."""
    return d - timedelta(days=d.weekday())


def fmt_week(ws):
    we = ws + timedelta(days=6)
    if ws.month == we.month:
        return f"Semaine du {ws.day} au {we.day} {MOIS_FR[we.month]} {we.year}"
    if ws.year == we.year:
        return (f"Semaine du {ws.day} {MOIS_FR[ws.month]} "
                f"au {we.day} {MOIS_FR[we.month]} {we.year}")
    return (f"Semaine du {ws.day} {MOIS_FR[ws.month]} {ws.year} "
            f"au {we.day} {MOIS_FR[we.month]} {we.year}")


def fmt_day(d):
    return f"{JOURS_FR[d.weekday()]} {d.day} {MOIS_FR[d.month]} {d.year}"


# ───────────────────────────────────────────────────────────────────────
#  Lecture du classeur
# ───────────────────────────────────────────────────────────────────────
def cell_rgb(ws, row, col):
    """Couleur de remplissage d'une cellule, en 'RRGGBB' majuscule, ou None."""
    c = ws.cell(row=row, column=col)
    if c.fill and c.fill.patternType == 'solid':
        fg = c.fill.fgColor
        if fg is not None and fg.type == 'rgb' and fg.rgb:
            rgb = str(fg.rgb).upper()
            return rgb[-6:]          # ignore le canal alpha éventuel
    return None


def read_concerts(path, sheet_name=None, rose='E0C2CD', vert='F6F9D4'):
    """Retourne la liste des concerts, chacun avec son niveau de signalement."""
    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Feuille « {sheet_name} » absente. "
                     f"Feuilles disponibles : {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        # priorité à « À venir », sinon la première feuille non vide
        ws = None
        for cand in ('À venir', 'A venir'):
            if cand in wb.sheetnames:
                ws = wb[cand]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

    rose = rose.upper().lstrip('#')[-6:]
    vert = vert.upper().lstrip('#')[-6:]

    concerts = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        jour, heure, salle, oeuvre, interp, comm, typ, tarifs = vals

        if not any([jour, salle, oeuvre]):
            continue

        rgb_oeuvre = cell_rgb(ws, r, 4)
        ligne = [cell_rgb(ws, r, c) for c in range(1, 9)]

        if rgb_oeuvre == rose:
            niveau = 'haut'                       # bordeaux
        elif rose in ligne or vert in ligne:
            niveau = 'signale'                    # vert
        else:
            niveau = 'normal'

        concerts.append({
            'jour':   str(jour).strip()   if jour   else '',
            'heure':  str(heure).strip()  if heure  else '',
            'salle':  str(salle).strip()  if salle  else '',
            'oeuvre': str(oeuvre).strip() if oeuvre else '',
            'interp': str(interp).strip() if interp else '',
            'comm':   str(comm).strip()   if comm   else '',
            'type':   str(typ).strip()    if typ    else '',
            'tarifs': str(tarifs).strip() if tarifs else '',
            'niveau': niveau,
        })

    return concerts, ws.title


# ───────────────────────────────────────────────────────────────────────
#  Styles
# ───────────────────────────────────────────────────────────────────────
def make_styles():
    S, B = FONT['SERIF'], FONT['SERIF-B']
    I = FONT['SERIF-I']
    SS, SSB = FONT['SANS'], FONT['SANS-B']
    s = {}

    s['cv_sur']   = ParagraphStyle('cv_sur', fontName=SS, fontSize=10, leading=14,
                                   textColor=GRIS_M, alignment=TA_CENTER, spaceAfter=5*mm)
    s['cv_titre'] = ParagraphStyle('cv_titre', fontName=B, fontSize=26, leading=31,
                                   textColor=GRIS_F, alignment=TA_CENTER, spaceAfter=3*mm)
    s['cv_sous']  = ParagraphStyle('cv_sous', fontName=SS, fontSize=10, leading=14,
                                   textColor=GRIS_M, alignment=TA_CENTER, spaceAfter=5*mm)
    s['url']      = ParagraphStyle('url', fontName=SS, fontSize=7.5, leading=10,
                                   textColor=OCRE, alignment=TA_LEFT, spaceAfter=6*mm)

    s['mois']     = ParagraphStyle('mois', fontName=B, fontSize=42, leading=48,
                                   textColor=GRIS_F, alignment=TA_CENTER)
    s['annee']    = ParagraphStyle('annee', fontName=SS, fontSize=16, leading=20,
                                   textColor=GRIS_M, alignment=TA_CENTER)

    s['sem']      = ParagraphStyle('sem', fontName=B, fontSize=18, leading=23,
                                   textColor=GRIS_F, alignment=TA_CENTER, spaceAfter=8*mm)
    s['sem_rub']  = ParagraphStyle('sem_rub', fontName=SSB, fontSize=7.5, leading=10,
                                   textColor=GRIS_M, alignment=TA_CENTER, spaceAfter=4*mm)
    s['sem_haut'] = ParagraphStyle('sem_haut', fontName=B, fontSize=9, leading=13,
                                   textColor=BORD, alignment=TA_CENTER, spaceAfter=1*mm)
    s['sem_sig']  = ParagraphStyle('sem_sig', fontName=B, fontSize=9, leading=13,
                                   textColor=VERT, alignment=TA_CENTER, spaceAfter=1*mm)
    s['sem_sub']  = ParagraphStyle('sem_sub', fontName=I, fontSize=7.5, leading=10,
                                   textColor=GRIS_M, alignment=TA_CENTER, spaceAfter=3*mm)

    s['jour']     = ParagraphStyle('jour', fontName=B, fontSize=11, leading=15,
                                   textColor=GRIS_F, spaceAfter=2*mm)
    s['heure']    = ParagraphStyle('heure', fontName=SS, fontSize=7.5, leading=10,
                                   textColor=GRIS_M, spaceBefore=1.5*mm)

    for key, col in (('o_normal', GRIS_F), ('o_haut', BORD), ('o_sig', VERT)):
        s[key] = ParagraphStyle(key, fontName=B, fontSize=9.5, leading=13,
                                textColor=col, spaceAfter=1.5*mm)

    s['salle']  = ParagraphStyle('salle', fontName=I, fontSize=8.5, leading=11.5,
                                 textColor=GRIS_M, spaceAfter=1*mm)
    s['interp'] = ParagraphStyle('interp', fontName=S, fontSize=8.5, leading=12,
                                 textColor=GRIS_F, spaceAfter=1.5*mm)
    s['comm']   = ParagraphStyle('comm', fontName=I, fontSize=8, leading=11,
                                 textColor=GRIS_M, spaceAfter=1*mm)
    s['meta']   = ParagraphStyle('meta', fontName=SS, fontSize=7.5, leading=10,
                                 textColor=GRIS_L)

    s['leg_t']  = ParagraphStyle('leg_t', fontName=SSB, fontSize=8, leading=11,
                                 textColor=GRIS_M, alignment=TA_CENTER, spaceAfter=4*mm)
    s['leg_i']  = ParagraphStyle('leg_i', fontName=SS, fontSize=8.5, leading=12,
                                 textColor=GRIS_F)
    return s


def esc(txt):
    """Échappe les caractères réservés du mini-langage XML de reportlab."""
    return (str(txt).replace('&', '&amp;')
                    .replace('<', '&lt;')
                    .replace('>', '&gt;'))


# ───────────────────────────────────────────────────────────────────────
#  Fonds de page
# ───────────────────────────────────────────────────────────────────────
def _fond(canvas, bg, filet_color, filet_w):
    canvas.saveState()
    canvas.setFillColor(bg)
    canvas.rect(0, 0, W, H, fill=1, stroke=0)
    canvas.setFillColor(filet_color)
    canvas.rect(10*mm, 0, filet_w, H, fill=1, stroke=0)
    canvas.restoreState()


def bg_cover(canvas, doc):
    _fond(canvas, CREME, OCRE, 4)


def bg_mois(canvas, doc):
    _fond(canvas, CREME_D, OCRE, 4)


def bg_semaine(canvas, doc):
    _fond(canvas, CREME, OCRE_L, 4)


def bg_content(canvas, doc):
    _fond(canvas, CREME, GRIS_L, 0.8)
    canvas.saveState()
    canvas.setFont(FONT['SANS'], 7)
    canvas.setFillColor(GRIS_L)
    canvas.drawCentredString(W / 2, 11*mm, str(doc.page - 1))
    canvas.restoreState()


# ───────────────────────────────────────────────────────────────────────
#  Blocs de contenu
# ───────────────────────────────────────────────────────────────────────
def build_cover(st, titre, sous_titre, surtitre, logo_path, url):
    story = []

    if logo_path and not os.path.isfile(logo_path):
        print(f"! Logo introuvable, ignoré : {logo_path}", file=sys.stderr)

    if logo_path and os.path.isfile(logo_path):
        img = Image(logo_path, width=38*mm, height=28*mm)
        img.hAlign = 'LEFT'
        story.append(img)
        story.append(Spacer(1, 2*mm))
        if url:
            story.append(Paragraph(
                f'<link href="{esc(url)}" color="#B07A20">{esc(url)}</link>', st['url']))
        story.append(Spacer(1, 6*mm))
    elif url:
        story.append(Paragraph(
            f'<link href="{esc(url)}" color="#B07A20">{esc(url)}</link>', st['url']))
        story.append(Spacer(1, 6*mm))
    else:
        story.append(Spacer(1, 42*mm))

    story.append(HRFlowable(width='100%', thickness=1.5, color=OCRE, spaceAfter=7*mm))
    if surtitre:
        story.append(Paragraph(esc(surtitre).upper(), st['cv_sur']))
        story.append(Spacer(1, 2*mm))
    story.append(Paragraph(esc(titre), st['cv_titre']))
    story.append(Spacer(1, 3*mm))
    if sous_titre:
        story.append(Paragraph(esc(sous_titre), st['cv_sous']))
    story.append(HRFlowable(width='100%', thickness=1.5, color=OCRE, spaceBefore=7*mm))

    story.append(Spacer(1, 40*mm))
    story.append(HRFlowable(width='60%', thickness=0.5, color=GRIS_L, spaceAfter=6*mm))
    story.append(Paragraph("CODES DE SIGNALEMENT", st['leg_t']))
    for bar, bg, label in ((BORD, BORD_L, "Hautement signalé — à ne pas manquer"),
                           (VERT, VERT_L, "Signalé — programme ou interprètes attrayants")):
        t = Table([[Paragraph(
            f'<font color="#{bar.hexval()[2:]}"><b>▌</b></font>  {label}', st['leg_i'])]],
            colWidths=[120*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), bg),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 2.5*mm))
    story.append(HRFlowable(width='60%', thickness=0.5, color=GRIS_L, spaceBefore=5*mm))
    return story


def build_mois(nom, annee, st):
    return [
        Spacer(1, H * 0.35 - MT),
        HRFlowable(width='80%', thickness=1.5, color=OCRE, spaceAfter=8*mm),
        Paragraph(nom.upper(), st['mois']),
        Spacer(1, 4*mm),
        Paragraph(str(annee), st['annee']),
        HRFlowable(width='80%', thickness=1.5, color=OCRE, spaceBefore=8*mm),
    ]


def _exergues(concerts, style_titre, style_sub):
    """Un titre par œuvre, avec toutes ses dates regroupées en sous-titre."""
    ordre, dates = [], defaultdict(list)
    for c in concerts:
        t = c['oeuvre']
        if t not in dates:
            ordre.append(t)
        d = parse_date(c['jour'])
        libelle = f"{JOURS_FR[d.weekday()]} {d.day}" if d else ''
        sub = ' · '.join(x for x in (libelle, c['salle']) if x)
        if sub and sub not in dates[t]:
            dates[t].append(sub)
    out = []
    for t in ordre:
        out.append(Paragraph(esc(t), style_titre))
        if dates[t]:
            out.append(Paragraph(esc(' — '.join(dates[t])), style_sub))
    return out


def build_semaine(ws_date, concerts_sem, st):
    story = [
        Spacer(1, 22*mm),
        HRFlowable(width='100%', thickness=1, color=OCRE, spaceAfter=6*mm),
        Paragraph(fmt_week(ws_date).upper(), st['sem']),
        HRFlowable(width='100%', thickness=1, color=OCRE, spaceAfter=8*mm),
    ]
    hauts   = [c for c in concerts_sem if c['niveau'] == 'haut']
    signales = [c for c in concerts_sem if c['niveau'] == 'signale']

    if hauts:
        story.append(Paragraph("HAUTEMENT SIGNALÉ", st['sem_rub']))
        story.append(Spacer(1, 2*mm))
        story.extend(_exergues(hauts, st['sem_haut'], st['sem_sub']))
    if hauts and signales:
        story.append(Spacer(1, 4*mm))
        story.append(HRFlowable(width='50%', thickness=0.4, color=GRIS_L, spaceAfter=4*mm))
    if signales:
        story.append(Paragraph("SIGNALÉ", st['sem_rub']))
        story.append(Spacer(1, 2*mm))
        story.extend(_exergues(signales, st['sem_sig'], st['sem_sub']))
    return story


def build_jour(d, st):
    return [
        HRFlowable(width='100%', thickness=2, color=OCRE, spaceAfter=1*mm),
        Paragraph(fmt_day(d).upper(), st['jour']),
        HRFlowable(width='100%', thickness=0.4, color=GRIS_L, spaceAfter=4*mm),
    ]


def build_concert(c, st, avail_w):
    niveau = c['niveau']
    if niveau == 'haut':
        style, bar, bg = st['o_haut'], BORD, BORD_L
    elif niveau == 'signale':
        style, bar, bg = st['o_sig'], VERT, VERT_L
    else:
        style, bar, bg = st['o_normal'], None, None

    droite = []
    marque = f'<font color="#{bar.hexval()[2:]}">▌ </font>' if bar else ''
    droite.append(Paragraph(marque + esc(c['oeuvre'] or '—'), style))
    if c['salle']:
        droite.append(Paragraph(esc(c['salle']), st['salle']))
    if c['interp']:
        droite.append(Paragraph(esc(c['interp']), st['interp']))
    if c['comm']:
        droite.append(Paragraph('[' + esc(c['comm']) + ']', st['comm']))
    meta = ' · '.join(x for x in (c['type'], c['tarifs']) if x)
    if meta:
        droite.append(Paragraph(esc(meta), st['meta']))

    gauche = [Paragraph(esc(c['heure']), st['heure'])] if c['heure'] else []

    HEURE_W = 16 * mm
    style_tbl = [
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',   (0, 0), (0, 0), 0),
        ('RIGHTPADDING',  (0, 0), (0, 0), 3*mm),
        ('LEFTPADDING',   (1, 0), (1, 0), 0),
        ('RIGHTPADDING',  (1, 0), (1, 0), 0),
        ('TOPPADDING',    (0, 0), (-1, -1), 2*mm),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3*mm),
    ]
    if bg:
        style_tbl += [
            ('BACKGROUND',    (0, 0), (-1, -1), bg),
            ('LEFTPADDING',   (0, 0), (0, 0), 3*mm),
            ('LEFTPADDING',   (1, 0), (1, 0), 3*mm),
            ('RIGHTPADDING',  (1, 0), (1, 0), 3*mm),
            ('TOPPADDING',    (0, 0), (-1, -1), 3*mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4*mm),
        ]
    t = Table([[gauche, droite]], colWidths=[HEURE_W, avail_w - HEURE_W])
    t.setStyle(TableStyle(style_tbl))
    return [t, HRFlowable(width='100%', thickness=0.4, color=GRIS_L, spaceAfter=3*mm)]


# ───────────────────────────────────────────────────────────────────────
#  Assemblage du document
# ───────────────────────────────────────────────────────────────────────
def build_pdf(concerts, sortie, titre, sous_titre, surtitre, logo, url):
    st = make_styles()
    avail_w = W - ML - MR

    valides = [c for c in concerts if parse_date(c['jour'])]
    valides.sort(key=lambda c: (c['jour'][:10], c['heure'] or ''))
    if not valides:
        sys.exit("Aucune ligne exploitable : vérifiez le format des dates "
                 "en colonne A (attendu « AAAA-MM-JJ »).")

    par_semaine = defaultdict(list)
    for c in valides:
        par_semaine[week_start(parse_date(c['jour']))].append(c)

    # type de fond, page par page
    types = ['cover']

    def on_page(canvas, doc):
        t = types[doc.page - 1] if doc.page <= len(types) else 'content'
        {'cover': bg_cover, 'mois': bg_mois,
         'semaine': bg_semaine}.get(t, bg_content)(canvas, doc)

    story = build_cover(st, titre, sous_titre, surtitre, logo, url)
    story.append(PageBreak())

    mois_cur = semaine_cur = jour_cur = None
    apres_saut = True     # la dernière opération était-elle un saut de page ?

    for c in valides:
        d = parse_date(c['jour'])
        mois = (d.year, d.month)
        sem = week_start(d)

        if mois != mois_cur:
            mois_cur, semaine_cur, jour_cur = mois, None, None
            if apres_saut:
                types[-1] = 'mois'
            else:
                types.append('mois')
                story.append(PageBreak())
            story.extend(build_mois(MOIS_FR[mois[1]], mois[0], st))
            types.append('semaine')
            story.append(PageBreak())
            apres_saut = True

        if sem != semaine_cur:
            semaine_cur, jour_cur = sem, None
            if apres_saut:
                types[-1] = 'semaine'
            else:
                types.append('semaine')
                story.append(PageBreak())
            story.extend(build_semaine(sem, par_semaine[sem], st))
            types.append('content')
            story.append(PageBreak())
            apres_saut = True

        if d != jour_cur:
            jour_cur = d
            if apres_saut:
                types[-1] = 'content'
            else:
                types.append('content')
                story.append(PageBreak())
            story.extend(build_jour(d, st))
            apres_saut = False

        story.extend(build_concert(c, st, avail_w))
        apres_saut = False

    doc = SimpleDocTemplate(
        sortie, pagesize=A4,
        leftMargin=ML, rightMargin=MR, topMargin=MT, bottomMargin=MB,
        title=titre, author=surtitre or '',
    )
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return doc.page, len(valides)


# ───────────────────────────────────────────────────────────────────────
#  Ligne de commande
# ───────────────────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(
        description="Transforme un relevé de concerts .xlsx en brochure PDF.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Exemple :\n"
               "  python3 brochure_concerts.py concerts_2026-27.xlsx --logo cello.jpg")
    p.add_argument('xlsx', help="fichier .xlsx source")
    p.add_argument('-o', '--sortie', help="PDF de sortie (défaut : même nom en .pdf)")
    p.add_argument('--feuille', help="nom de la feuille (défaut : « À venir », sinon la première)")
    p.add_argument('--logo', help="image de couverture (jpg/png)")
    p.add_argument('--url', default='http://carnetsol.fr/css/',
                   help="lien cliquable en couverture (défaut : %(default)s ; "
                        "chaîne vide pour l'omettre)")
    p.add_argument('--titre', default="Concerts & Spectacles musicaux")
    p.add_argument('--surtitre', default="Agenda de Carnets sur sol")
    p.add_argument('--sous-titre', dest='sous_titre',
                   default="Saisons 2025-6 et 2026-7 · À Paris & en Francilie")
    p.add_argument('--du', help="ne garder que les concerts à partir de cette date (AAAA-MM-JJ)")
    p.add_argument('--au', help="ne garder que les concerts jusqu'à cette date (AAAA-MM-JJ)")
    p.add_argument('--rose', default='E0C2CD',
                   help="teinte « hautement signalé » (défaut : %(default)s)")
    p.add_argument('--vert', default='F6F9D4',
                   help="teinte « signalé » (défaut : %(default)s)")
    p.add_argument('--polices', help="dossier contenant les .ttf DejaVu")
    args = p.parse_args()

    if not os.path.isfile(args.xlsx):
        sys.exit(f"Fichier introuvable : {args.xlsx}")

    sortie = args.sortie or os.path.splitext(args.xlsx)[0] + '.pdf'
    setup_fonts(args.polices)

    concerts, feuille = read_concerts(args.xlsx, args.feuille, args.rose, args.vert)
    print(f"Feuille lue : « {feuille} » — {len(concerts)} lignes")

    if args.du or args.au:
        debut = parse_date(args.du) if args.du else None
        fin = parse_date(args.au) if args.au else None
        avant = len(concerts)
        gardes = []
        for c in concerts:
            d = parse_date(c['jour'])
            if d and (debut is None or d >= debut) and (fin is None or d <= fin):
                gardes.append(c)
        concerts = gardes
        print(f"Filtre de dates : {len(concerts)} lignes retenues sur {avant}")

    n_haut = sum(1 for c in concerts if c['niveau'] == 'haut')
    n_sig = sum(1 for c in concerts if c['niveau'] == 'signale')
    print(f"Signalements : {n_haut} hautement signalés, {n_sig} signalés")

    pages, retenus = build_pdf(
        concerts, sortie,
        args.titre, args.sous_titre, args.surtitre,
        args.logo, args.url)

    print(f"✓ {sortie} — {pages} pages, {retenus} concerts")


if __name__ == '__main__':
    main()
